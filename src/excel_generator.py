"""
excel_generator.py - Generacion del Excel maestro con metadata
"""

import pandas as pd
from pathlib import Path
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from loguru import logger


class ExcelGenerator:
    """Genera y formatea el archivo Excel maestro."""

    def __init__(self):
        """Inicializa el generador de Excel."""
        self.columnas = [
            'ID Publicacion',
            'Fecha Publicacion',
            'Pagina',
            'Titulo',
            'Hashtags',
            'Me Gusta',
            'Me Enoja',
            'Me Encanta',
            'Total Reacciones',
            'Comentarios',
            'Compartidos',
            'Total Interacciones',
            'URL Original',
            'Carpeta Local',
            'Imagen',
            'Fecha Scraping'
        ]

    def generar_excel(self, datos: List[Dict[str, Any]], ruta_salida: Path) -> bool:
        """
        Genera el archivo Excel maestro.

        Args:
            datos: Lista de diccionarios con datos de publicaciones
            ruta_salida: Ruta donde guardar el Excel

        Returns:
            True si se genero correctamente
        """
        if not datos:
            logger.warning("No hay datos para generar Excel")
            return False

        try:
            # Preparar datos para DataFrame
            filas = []
            for pub in datos:
                reacciones = pub.get('reacciones', {})

                fila = {
                    'ID Publicacion': pub.get('id_publicacion', 'N/A'),
                    'Fecha Publicacion': pub.get('fecha_publicacion', 'N/A'),
                    'Pagina': pub.get('pagina', 'N/A'),
                    'Titulo': self._truncar_texto(pub.get('titulo', ''), 100),
                    'Hashtags': ', '.join(pub.get('hashtags', [])),
                    'Me Gusta': reacciones.get('me_gusta', 0),
                    'Me Enoja': reacciones.get('me_enoja', 0),
                    'Me Encanta': reacciones.get('me_encanta', 0),
                    'Total Reacciones': reacciones.get('total', 0),
                    'Comentarios': pub.get('comentarios', 0),
                    'Compartidos': pub.get('compartidos', 0),
                    'Total Interacciones': (
                        reacciones.get('total', 0) +
                        pub.get('comentarios', 0) +
                        pub.get('compartidos', 0)
                    ),
                    'URL Original': pub.get('url_publicacion', ''),
                    'Carpeta Local': pub.get('carpeta', ''),
                    'Imagen': pub.get('imagen_local', ''),
                    'Fecha Scraping': pub.get('fecha_scraping', '')
                }
                filas.append(fila)

            # Crear DataFrame
            df = pd.DataFrame(filas, columns=self.columnas)

            # Crear Excel con formato
            self._crear_excel_formateado(df, ruta_salida)

            logger.info(f"Excel generado: {ruta_salida}")
            return True

        except Exception as e:
            logger.error(f"Error generando Excel: {e}")
            return False

    def _crear_excel_formateado(self, df: pd.DataFrame, ruta: Path):
        """
        Crea un Excel con formato profesional.

        Args:
            df: DataFrame con los datos
            ruta: Ruta de salida
        """
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Publicaciones Facebook"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        cell_alignment = Alignment(vertical="center", wrap_text=True)
        number_alignment = Alignment(horizontal="right", vertical="center")

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Escribir headers
        for col_idx, columna in enumerate(self.columnas, 1):
            cell = ws.cell(row=1, column=col_idx, value=columna)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Escribir datos
        for row_idx, fila in enumerate(df.values, 2):
            for col_idx, valor in enumerate(fila, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=valor)
                cell.border = thin_border

                # Alineacion segun tipo de dato
                if isinstance(valor, (int, float)):
                    cell.alignment = number_alignment
                else:
                    cell.alignment = cell_alignment

        # Ajustar ancho de columnas
        anchos = {
            'ID Publicacion': 18,
            'Fecha Publicacion': 15,
            'Pagina': 20,
            'Titulo': 50,
            'Hashtags': 25,
            'Me Gusta': 10,
            'Me Enoja': 10,
            'Me Encanta': 10,
            'Total Reacciones': 12,
            'Comentarios': 12,
            'Compartidos': 12,
            'Total Interacciones': 15,
            'URL Original': 40,
            'Carpeta Local': 35,
            'Imagen': 15,
            'Fecha Scraping': 18
        }

        for col_idx, columna in enumerate(self.columnas, 1):
            letra = get_column_letter(col_idx)
            ws.column_dimensions[letra].width = anchos.get(columna, 15)

        # Congelar primera fila
        ws.freeze_panes = 'A2'

        # Agregar auto-filtro
        ws.auto_filter.ref = ws.dimensions

        # Altura de la fila de encabezado
        ws.row_dimensions[1].height = 30

        # Guardar
        wb.save(ruta)
        wb.close()

    def _truncar_texto(self, texto: str, max_chars: int) -> str:
        """
        Trunca un texto a un maximo de caracteres.

        Args:
            texto: Texto original
            max_chars: Maximo de caracteres

        Returns:
            Texto truncado
        """
        if not texto:
            return ""

        if len(texto) <= max_chars:
            return texto

        return texto[:max_chars - 3] + "..."

    def agregar_hoja_resumen(self, ruta_excel: Path, estadisticas: Dict[str, Any]):
        """
        Agrega una hoja de resumen al Excel existente.

        Args:
            ruta_excel: Ruta al Excel
            estadisticas: Diccionario con estadisticas
        """
        try:
            from openpyxl import load_workbook

            wb = load_workbook(ruta_excel)
            ws = wb.create_sheet("Resumen", 0)

            # Estilos
            titulo_font = Font(bold=True, size=14)
            valor_font = Font(size=12)

            # Datos del resumen
            resumen = [
                ("RESUMEN DE SCRAPING", ""),
                ("", ""),
                ("Total de publicaciones", estadisticas.get('total', 0)),
                ("Publicaciones exitosas", estadisticas.get('exitosas', 0)),
                ("Publicaciones fallidas", estadisticas.get('fallidas', 0)),
                ("Tasa de exito", f"{estadisticas.get('tasa_exito', 0):.1f}%"),
                ("", ""),
                ("Imagenes descargadas", estadisticas.get('imagenes', 0)),
                ("Espacio utilizado", f"{estadisticas.get('espacio_mb', 0):.2f} MB"),
                ("", ""),
                ("Fecha de ejecucion", estadisticas.get('fecha', '')),
                ("Duracion", estadisticas.get('duracion', ''))
            ]

            for row_idx, (etiqueta, valor) in enumerate(resumen, 1):
                ws.cell(row=row_idx, column=1, value=etiqueta).font = titulo_font if row_idx == 1 else valor_font
                ws.cell(row=row_idx, column=2, value=valor).font = valor_font

            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 20

            wb.save(ruta_excel)
            wb.close()

        except Exception as e:
            logger.error(f"Error agregando hoja de resumen: {e}")

    def exportar_csv(self, datos: List[Dict[str, Any]], ruta_salida: Path) -> bool:
        """
        Exporta los datos a CSV como alternativa.

        Args:
            datos: Lista de datos
            ruta_salida: Ruta de salida

        Returns:
            True si se exporto correctamente
        """
        try:
            filas = []
            for pub in datos:
                reacciones = pub.get('reacciones', {})
                fila = {
                    'id': pub.get('id_publicacion', ''),
                    'fecha': pub.get('fecha_publicacion', ''),
                    'pagina': pub.get('pagina', ''),
                    'titulo': pub.get('titulo', ''),
                    'hashtags': '|'.join(pub.get('hashtags', [])),
                    'reacciones': reacciones.get('total', 0),
                    'comentarios': pub.get('comentarios', 0),
                    'compartidos': pub.get('compartidos', 0),
                    'url': pub.get('url_publicacion', ''),
                    'carpeta': pub.get('carpeta', '')
                }
                filas.append(fila)

            df = pd.DataFrame(filas)
            df.to_csv(ruta_salida, index=False, encoding='utf-8-sig')
            return True

        except Exception as e:
            logger.error(f"Error exportando CSV: {e}")
            return False
